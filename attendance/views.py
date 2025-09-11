import base64
import io
import os
from datetime import date, datetime
import cv2
import numpy as np
from pyzbar import pyzbar
import qrcode
from openpyxl import Workbook, load_workbook
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.conf import settings
from .models import AttendanceRecord

def qr_card(request):
    """Render a card and, on POST, generate a date-only QR."""
    ctx = {
        "name": "M.Abinaya",
        "department": "Bsc.CS",
        "year": "3rd year",
        "date": "",
        "qr_b64": None,
        "in_time": "Not marked",
        "out_time": "Not marked",
    }

    if request.method == "POST":
        today = date.today().isoformat()       # e.g., 2025-09-07
        
        # Create QR content with all student information and date
        qr_content = f"""Name: M.Abinaya
Department: Bsc.CS
Year: 3rd year
Date: {today}"""
        
        # Generate QR code with all information
        img = qrcode.make(qr_content)

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        ctx["qr_b64"] = base64.b64encode(buf.getvalue()).decode("utf-8")
        ctx["date"] = today

    return render(request, "attendance/qr_card.html", ctx)


def create_excel_file():
    """Create or update the Excel file with attendance data"""
    excel_file_path = os.path.join(settings.BASE_DIR, 'attendance_data', 'attendance_records.xlsx')
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)
    
    try:
        # Try to load existing workbook
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Records"
        
        # Add headers
        headers = ["Student Name", "Department", "Year", "Scanned By", "Scanner Location", "Scanner Device", "Date", "Time", "Full Timestamp"]
        worksheet.append(headers)
    
    return workbook, worksheet, excel_file_path


@csrf_exempt
def scan_qr_code(request):
    """Handle QR code scanning from uploaded image"""
    if request.method == 'POST':
        try:
            # Get uploaded file
            uploaded_file = request.FILES.get('qr_image')
            if not uploaded_file:
                return JsonResponse({'error': 'No image uploaded'}, status=400)
            
            # Read image data
            image_data = uploaded_file.read()
            nparr = np.frombuffer(image_data, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            # Decode QR codes
            qr_codes = pyzbar.decode(image)
            
            if not qr_codes:
                return JsonResponse({'error': 'No QR code found in image'}, status=400)
            
            # Process the first QR code found
            qr_data = qr_codes[0].data.decode('utf-8')
            
            # Parse QR code data
            lines = qr_data.strip().split('\n')
            student_info = {}
            
            for line in lines:
                if ':' in line:
                    key, value = line.split(':', 1)
                    student_info[key.strip().lower()] = value.strip()
            
            # Extract information
            name = student_info.get('name', 'Unknown')
            department = student_info.get('department', 'Unknown')
            year = student_info.get('year', 'Unknown')
            
            # Get scanner information from request
            scanner_name = request.POST.get('scanner_name', 'Unknown Scanner')
            scanner_location = request.POST.get('scanner_location', '')
            scanner_device = request.META.get('HTTP_USER_AGENT', 'Unknown Device')[:100]
            
            # Save to database (avoid duplicates for same day)
            record, created = AttendanceRecord.objects.get_or_create(
                name=name,
                scan_date=date.today(),
                defaults={
                    'department': department,
                    'year': year,
                    'scanner_name': scanner_name,
                    'scanner_location': scanner_location,
                    'scanner_device': scanner_device
                }
            )
            
            # Update Excel file
            workbook, worksheet, excel_file_path = create_excel_file()
            
            print(f"DEBUG: Excel file path: {excel_file_path}")
            print(f"DEBUG: Record created: {created}")
            print(f"DEBUG: Scanner name: {scanner_name}")
            
            if created:  # Only add to Excel if it's a new record for today
                current_time = datetime.now()
                new_row = [
                    name,  # Student Name
                    department,  # Department
                    year,  # Year
                    scanner_name,  # Scanned By
                    scanner_location or 'Not specified',  # Scanner Location
                    scanner_device[:50],  # Scanner Device (truncated)
                    current_time.strftime('%Y-%m-%d'),  # Date
                    current_time.strftime('%H:%M:%S'),  # Time
                    current_time.strftime('%Y-%m-%d %H:%M:%S')  # Full Timestamp
                ]
                worksheet.append(new_row)
                workbook.save(excel_file_path)
                print(f"DEBUG: Excel file saved successfully at {excel_file_path}")
            else:
                print("DEBUG: Record already exists for today, not adding to Excel")
            
            return JsonResponse({
                'success': True,
                'data': {
                    'name': name,
                    'department': department,
                    'year': year,
                    'scanner_name': scanner_name,
                    'scanner_location': scanner_location,
                    'date': record.scan_date.strftime('%Y-%m-%d'),
                    'time': record.scan_time.strftime('%H:%M:%S'),
                    'already_scanned': not created
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Error processing QR code: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Only POST method allowed'}, status=405)


def download_excel(request):
    """Download the attendance Excel file"""
    excel_file_path = os.path.join(settings.BASE_DIR, 'attendance_data', 'attendance_records.xlsx')
    
    if not os.path.exists(excel_file_path):
        # Create empty file if doesn't exist
        workbook, worksheet, excel_file_path = create_excel_file()
        workbook.save(excel_file_path)
    
    try:
        with open(excel_file_path, 'rb') as excel_file:
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="attendance_records_{date.today()}.xlsx"'
            return response
    except Exception as e:
        return JsonResponse({'error': f'Error downloading file: {str(e)}'}, status=500)


def scanner_page(request):
    """Render the QR code scanner page"""
    return render(request, 'attendance/scanner.html')


def view_records(request):
    """View all attendance records"""
    records = AttendanceRecord.objects.all().order_by('-created_at')
    return render(request, 'attendance/records.html', {'records': records})
